"""Export analysis results to Excel and PDF."""

from __future__ import annotations

import datetime as dt
from pathlib import Path
from typing import Dict, List

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.config import BLUE_THEME, METRIC_LABELS
from app.core.paths import get_reports_path, get_temp_charts_path
from app.models.entities import AnalysisPackage, MetricExplanation
from app.utils.numbers import format_currency, format_number


class ReportExporter:
    def __init__(self) -> None:
        self.reports_dir = get_reports_path(create_if_missing=True)
        self.temp_charts_dir = get_temp_charts_path(create_if_missing=True)

    def export_excel(
        self,
        analysis: AnalysisPackage,
        metric_keys: List[str],
        explanations: Dict[str, MetricExplanation],
        output_path: Path | None = None,
    ) -> Path:
        output_path = output_path or self._default_output_path(analysis=analysis, extension="xlsx")

        metrics_df = analysis.metrics_dataframe(metric_keys)
        metrics_df = metrics_df.rename(
            columns={k: METRIC_LABELS.get(k, k.replace("_", " ").title()) for k in metric_keys}
        )

        income_df, balance_df, cash_df = self._statement_dataframes(analysis=analysis)
        explanations_df = self._explanations_dataframe(metric_keys=metric_keys, explanations=explanations)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            metrics_df.to_excel(writer, index=False, sheet_name="Indicadores")
            income_df.to_excel(writer, index=False, sheet_name="EstadoResultados")
            balance_df.to_excel(writer, index=False, sheet_name="BalanceGeneral")
            cash_df.to_excel(writer, index=False, sheet_name="FlujoCaja")
            explanations_df.to_excel(writer, index=False, sheet_name="Explicaciones")

        workbook = load_workbook(output_path)
        chart_sheet = workbook.create_sheet("Graficas")

        row_anchor = 1
        for metric_key in metric_keys:
            chart_file = self._save_metric_chart(analysis=analysis, metric_key=metric_key)
            chart_sheet[f"A{row_anchor}"] = METRIC_LABELS.get(metric_key, metric_key)
            image = OpenpyxlImage(str(chart_file))
            image.width = 760
            image.height = 320
            chart_sheet.add_image(image, f"A{row_anchor + 1}")
            row_anchor += 20

        workbook.save(output_path)
        return output_path

    def export_pdf(
        self,
        analysis: AnalysisPackage,
        metric_keys: List[str],
        explanations: Dict[str, MetricExplanation],
        output_path: Path | None = None,
    ) -> Path:
        output_path = output_path or self._default_output_path(analysis=analysis, extension="pdf")

        doc = SimpleDocTemplate(str(output_path), pagesize=A4, rightMargin=1.5 * cm, leftMargin=1.5 * cm)
        styles = getSampleStyleSheet()
        story = []

        story.append(Paragraph("Reporte financiero - Analizador de Empresas (Supersociedades)", styles["Title"]))
        story.append(
            Paragraph(
                f"Empresa: {analysis.company.razon_social} | NIT: {analysis.company.nit}",
                styles["Normal"],
            )
        )
        story.append(Spacer(1, 0.4 * cm))

        metrics_df = analysis.metrics_dataframe(metric_keys)
        table_data = [["Ano"] + [METRIC_LABELS.get(k, k) for k in metric_keys]]
        for _, row in metrics_df.sort_values("anio").iterrows():
            table_row = [str(int(row["anio"]))]
            for key in metric_keys:
                value = row.get(key)
                if key in {"dias_capital_trabajo", "z_altman"}:
                    table_row.append(format_number(value, 2))
                else:
                    table_row.append(format_currency(value))
            table_data.append(table_row)

        table = Table(table_data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BLUE_THEME["primary"])),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#A7C3E8")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 0.5 * cm))

        for metric_key in metric_keys:
            story.append(Paragraph(METRIC_LABELS.get(metric_key, metric_key), styles["Heading3"]))
            chart_file = self._save_metric_chart(analysis=analysis, metric_key=metric_key)
            story.append(Image(str(chart_file), width=17.5 * cm, height=6.8 * cm))

            exp = explanations.get(metric_key)
            if exp:
                story.append(Paragraph(f"Que significa: {exp.what_is}", styles["BodyText"]))
                story.append(Paragraph(f"Como interpretarlo: {exp.interpretation}", styles["BodyText"]))
                story.append(Paragraph(f"Senales positivas/negativas: {exp.signals}", styles["BodyText"]))
                story.append(Paragraph(f"Preguntas de negocio: {exp.business_questions}", styles["BodyText"]))
            story.append(Spacer(1, 0.3 * cm))

        warnings = analysis.warnings()
        if warnings:
            story.append(Paragraph("Advertencias de calidad de datos", styles["Heading3"]))
            for warning in warnings:
                story.append(Paragraph(f"- {warning}", styles["BodyText"]))

        doc.build(story)
        return output_path

    def _default_output_path(self, analysis: AnalysisPackage, extension: str) -> Path:
        timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        name = f"reporte_{analysis.company.nit}_{timestamp}.{extension}"
        return self.reports_dir / name

    def _statement_dataframes(self, analysis: AnalysisPackage) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        income_rows = []
        balance_rows = []
        cash_rows = []

        for year in sorted(analysis.years):
            snapshot = analysis.snapshots[year]

            income_rows.append({"ano": year, **snapshot.income_statement})
            balance_rows.append({"ano": year, **snapshot.balance_sheet})
            cash_rows.append({"ano": year, **snapshot.cash_flow})

        return pd.DataFrame(income_rows), pd.DataFrame(balance_rows), pd.DataFrame(cash_rows)

    @staticmethod
    def _explanations_dataframe(
        metric_keys: List[str],
        explanations: Dict[str, MetricExplanation],
    ) -> pd.DataFrame:
        rows = []
        for metric_key in metric_keys:
            exp = explanations.get(metric_key)
            if not exp:
                continue
            rows.append(
                {
                    "indicador": METRIC_LABELS.get(metric_key, metric_key),
                    "que_significa": exp.what_is,
                    "como_interpretarlo": exp.interpretation,
                    "senales": exp.signals,
                    "preguntas_de_negocio": exp.business_questions,
                }
            )
        return pd.DataFrame(rows)

    def _save_metric_chart(self, analysis: AnalysisPackage, metric_key: str) -> Path:
        years = sorted(analysis.years)
        path = self.temp_charts_dir / f"{analysis.company.nit}_{metric_key}.png"

        fig, ax = plt.subplots(figsize=(9, 3.8))
        fig.patch.set_facecolor("white")

        if metric_key == "balance_general":
            activos = [analysis.snapshots[y].balance_sheet.get("activos_totales") for y in years]
            pasivos = [analysis.snapshots[y].balance_sheet.get("pasivos_totales") for y in years]
            patrimonio = [analysis.snapshots[y].balance_sheet.get("patrimonio_total") for y in years]

            activos = [float("nan") if v is None else v for v in activos]
            pasivos = [float("nan") if v is None else v for v in pasivos]
            patrimonio = [float("nan") if v is None else v for v in patrimonio]

            ax.plot(years, activos, marker="o", color="#0F4C81", label="Activos")
            ax.plot(years, pasivos, marker="o", color="#C53030", label="Pasivos")
            ax.plot(years, patrimonio, marker="o", color="#2F855A", label="Patrimonio")
            ax.set_ylabel("COP")
            ax.legend(loc="best")
        else:
            values = [analysis.snapshots[y].metrics.get(metric_key) for y in years]
            values = [0 if v is None else v for v in values]
            ax.bar(years, values, color="#2E7CBF")
            ax.set_ylabel("Valor")

        ax.set_title(METRIC_LABELS.get(metric_key, metric_key))
        ax.grid(axis="y", linestyle="--", alpha=0.3)
        ax.set_xlabel("Ano")

        fig.tight_layout()
        fig.savefig(path, dpi=160)
        plt.close(fig)
        return path
